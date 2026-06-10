"""
Agente de Higienização de Vendas da Porsche
============================================

Lê um arquivo .xlsx de vendas Porsche (bruto) e aplica as regras definidas em
schema.md para gerar um novo arquivo tratado, com cada coluna sanitizada
inserida logo após sua coluna de origem.

Uso:
    python sanitize_porsche.py [ARQUIVO_ENTRADA] [ARQUIVO_SAIDA]

Por padrão, usa INPUT_PATH e OUTPUT_PATH definidos abaixo.
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Valores padrão
# ---------------------------------------------------------------------------
INPUT_PATH  = Path("porsche_database.xlsx")
OUTPUT_PATH = Path("porsche_database_sanitized.xlsx")
INVALIDO    = "INVALIDO"


# ---------------------------------------------------------------------------
# Higienização de datas
# ---------------------------------------------------------------------------
MESES_PT = {
    "jan": 1, "janeiro": 1,
    "fev": 2, "fevereiro": 2,
    "mar": 3, "marco": 3, "março": 3,
    "abr": 4, "abril": 4,
    "mai": 5, "maio": 5,
    "jun": 6, "junho": 6,
    "jul": 7, "julho": 7,
    "ago": 8, "agosto": 8,
    "set": 9, "setembro": 9,
    "out": 10, "outubro": 10,
    "nov": 11, "novembro": 11,
    "dez": 12, "dezembro": 12,
}
MESES_EN = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}
MESES = {**MESES_PT, **MESES_EN}


def _data_segura(y: int, m: int, d: int) -> str:
    try:
        return dt.date(y, m, d).strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return INVALIDO


def sanitizar_data(valor: Any) -> str:
    if valor is None or valor == "":
        return INVALIDO

    if isinstance(valor, dt.datetime):
        return valor.strftime("%Y-%m-%d")
    if isinstance(valor, dt.date):
        return valor.strftime("%Y-%m-%d")

    # Número serial do Excel (dias desde 1900-01-01)
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        try:
            base = dt.date(1899, 12, 30)
            d = base + dt.timedelta(days=int(valor))
            return d.strftime("%Y-%m-%d")
        except Exception:
            return INVALIDO

    s = str(valor).strip()
    if not s:
        return INVALIDO

    # AAAA-MM-DD / AAAA/MM/DD / AAAA.MM.DD
    m = re.match(r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})$", s)
    if m:
        return _data_segura(int(m.group(1)), int(m.group(2)), int(m.group(3)))

    # MM/DD/AAAA ou MM-DD-AAAA
    m = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$", s)
    if m:
        return _data_segura(int(m.group(3)), int(m.group(1)), int(m.group(2)))

    # MM/DD/AA ou MM-DD-AA
    m = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{2})$", s)
    if m:
        yy = int(m.group(3))
        ano = 2000 + yy if yy < 70 else 1900 + yy
        return _data_segura(ano, int(m.group(1)), int(m.group(2)))

    # "Mês DD, AAAA" / "Mês DDº AAAA" / "Mês DD de AAAA"
    m = re.match(
        r"^([A-Za-záéíóúãõâêîôûàèìòùç]+)\.?\s+(\d{1,2})(?:st|nd|rd|th)?[,\s]+(\d{4})$",
        s, re.IGNORECASE,
    )
    if m:
        chave = m.group(1).lower().rstrip(".")
        if chave in MESES:
            return _data_segura(int(m.group(3)), MESES[chave], int(m.group(2)))

    return INVALIDO


# ---------------------------------------------------------------------------
# Higienização do modelo Porsche
# ---------------------------------------------------------------------------
MODELOS_CANONICOS = [
    "911 Carrera", "911 Carrera S", "911 Carrera GTS",
    "911 Turbo", "911 Turbo S",
    "911 GT3", "911 GT3 RS",
    "911 Dakar",
    "911 Targa 4", "911 Targa 4S",
    "718 Cayman", "718 Cayman S", "718 Cayman GT4 RS",
    "718 Boxster", "718 Boxster GTS",
    "718 Spyder RS",
    "Cayenne", "Cayenne S", "Cayenne Coupe", "Cayenne E-Hybrid",
    "Cayenne Turbo", "Cayenne Turbo GT",
    "Macan", "Macan S", "Macan T", "Macan GTS", "Macan Electric",
    "Panamera", "Panamera 4", "Panamera 4S",
    "Panamera Turbo", "Panamera Turbo S", "Panamera 4 E-Hybrid",
    "Taycan", "Taycan 4S", "Taycan GTS",
    "Taycan Turbo", "Taycan Turbo S", "Taycan Cross Turismo",
]

_MODELO_LOOKUP = {m.lower(): m for m in MODELOS_CANONICOS}


def _titulo_inteligente(s: str) -> str:
    """Title case preservando siglas como GTS, RS, E-Hybrid, etc."""
    manter_maiusculo = {"GT", "GTS", "RS", "GT3", "GT4", "S", "T", "4", "4S"}
    saida = []
    for p in s.split():
        if p.upper() in manter_maiusculo:
            saida.append(p.upper())
        elif "-" in p:
            saida.append("-".join(seg.capitalize() for seg in p.split("-")))
        else:
            saida.append(p.capitalize())
    return " ".join(saida)


def sanitizar_modelo(valor: Any) -> str:
    if valor is None:
        return INVALIDO
    s = " ".join(str(valor).split()).strip()
    if not s:
        return INVALIDO
    chave = s.lower()
    if chave in _MODELO_LOOKUP:
        return _MODELO_LOOKUP[chave]
    # Modelos desconhecidos: title case
    return _titulo_inteligente(s)


# ---------------------------------------------------------------------------
# Higienização do ano do modelo
# ---------------------------------------------------------------------------
PALAVRAS_PARA_NUM: dict[str, int] = {
    # Português
    "zero": 0, "um": 1, "dois": 2, "tres": 3, "três": 3, "quatro": 4,
    "cinco": 5, "seis": 6, "sete": 7, "oito": 8, "nove": 9,
    "dez": 10, "onze": 11, "doze": 12, "treze": 13,
    "quatorze": 14, "quinze": 15, "dezesseis": 16, "dezessete": 17,
    "dezoito": 18, "dezenove": 19, "vinte": 20, "trinta": 30,
    "quarenta": 40, "cinquenta": 50, "sessenta": 60, "setenta": 70,
    "oitenta": 80, "noventa": 90, "cem": 100, "mil": 1000,
    # Inglês
    "zero": 0, "one": 1, "two": 2, "three": 3, "four": 4,
    "five": 5, "six": 6, "seven": 7, "eight": 8, "nine": 9,
    "ten": 10, "eleven": 11, "twelve": 12, "thirteen": 13,
    "fourteen": 14, "fifteen": 15, "sixteen": 16, "seventeen": 17,
    "eighteen": 18, "nineteen": 19, "twenty": 20, "thirty": 30,
    "forty": 40, "fifty": 50, "sixty": 60, "seventy": 70,
    "eighty": 80, "ninety": 90, "hundred": 100, "thousand": 1000,
}


def _palavras_para_int(texto: str) -> int | None:
    tokens = re.findall(r"[a-záéíóúãõâêç]+", texto.lower())
    if not tokens:
        return None
    if any(t not in PALAVRAS_PARA_NUM for t in tokens):
        return None

    if all(PALAVRAS_PARA_NUM[t] < 100 for t in tokens):
        nums = [PALAVRAS_PARA_NUM[t] for t in tokens]
        grupos: list[int] = []
        i = 0
        while i < len(nums):
            if (i + 1 < len(nums) and nums[i] >= 20
                    and nums[i] % 10 == 0 and nums[i + 1] < 10):
                grupos.append(nums[i] + nums[i + 1])
                i += 2
            else:
                grupos.append(nums[i])
                i += 1
        if len(grupos) == 2:
            return grupos[0] * 100 + grupos[1]
        if len(grupos) == 1:
            return grupos[0]

    total, atual = 0, 0
    for t in tokens:
        v = PALAVRAS_PARA_NUM[t]
        if v in (100, 1000):
            atual = max(1, atual) * v
            if v == 1000:
                total += atual
                atual = 0
        else:
            atual += v
    return total + atual


def sanitizar_ano(valor: Any) -> str:
    if valor is None:
        return INVALIDO

    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        y = int(valor)
        return str(y) if 1990 <= y <= 2035 else INVALIDO

    s = str(valor).strip()
    if not s:
        return INVALIDO

    # 4 dígitos exatos
    if re.fullmatch(r"\d{4}", s):
        y = int(s)
        return str(y) if 1990 <= y <= 2035 else INVALIDO

    # "20-24", "20 24", "20.24"
    m = re.fullmatch(r"(\d{2})\s*[-\s.]\s*(\d{2})", s)
    if m:
        y = int(m.group(1) + m.group(2))
        return str(y) if 1990 <= y <= 2035 else INVALIDO

    # Texto por extenso
    if re.search(r"[A-Za-záéíóúãõ]", s):
        n = _palavras_para_int(s)
        if n is not None and 1990 <= n <= 2035:
            return str(n)

    return INVALIDO


# ---------------------------------------------------------------------------
# Higienização do preço de venda
# ---------------------------------------------------------------------------
def _parse_preco_textual(s: str) -> float | None:
    tokens = re.findall(r"[a-záéíóúãõ]+", s.lower())
    if not tokens:
        return None
    validos = set(PALAVRAS_PARA_NUM) | {"milhao", "milhão", "million"}
    if any(t not in validos for t in tokens):
        return None

    total, atual = 0, 0
    for t in tokens:
        if t in ("milhao", "milhão", "million"):
            atual = max(1, atual) * 1_000_000
            total += atual
            atual = 0
        elif t == "cem":
            atual = max(1, atual) * 100
        elif t in ("mil", "thousand"):
            atual = max(1, atual) * 1_000
            total += atual
            atual = 0
        elif t in PALAVRAS_PARA_NUM:
            atual += PALAVRAS_PARA_NUM[t]
    return float(total + atual)


def _parse_preco_numerico(num_str: str) -> float | None:
    s = num_str.strip()
    if not s:
        return None

    tem_ponto   = "." in s
    tem_virgula = "," in s

    if tem_ponto and tem_virgula:
        # Último separador é o decimal
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")   # europeu
        else:
            s = s.replace(",", "")                      # americano
    elif tem_virgula:
        ultimo = s.rsplit(",", 1)
        digitos_apos = re.sub(r"\D", "", ultimo[1])
        if s.count(",") > 1 or len(digitos_apos) != 2:
            s = s.replace(",", "")
        else:
            s = s.replace(",", ".")
    elif tem_ponto:
        ultimo = s.rsplit(".", 1)
        digitos_apos = re.sub(r"\D", "", ultimo[1])
        if s.count(".") > 1 or len(digitos_apos) != 2:
            s = s.replace(".", "")

    try:
        return float(s)
    except ValueError:
        return None


def sanitizar_preco(valor: Any) -> str:
    if valor is None or valor == "":
        return INVALIDO

    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        return f"{float(valor):.2f}"

    s = str(valor).strip()
    if not s:
        return INVALIDO

    limpo = s.replace("$", " ")
    limpo = re.sub(
        r"\b(usd|dollars?|dólares?|reais?|brl)\b", " ", limpo, flags=re.IGNORECASE
    )
    limpo = limpo.strip()

    # Apenas texto (ex: "oitenta e dois mil")
    if re.fullmatch(r"[A-Za-záéíóúãõâêç\s]+", limpo):
        n = _parse_preco_textual(limpo)
        return f"{n:.2f}" if n is not None else INVALIDO

    # Sufixo "k" (mil)
    multiplicador = 1.0
    m = re.search(r"([0-9.,]+)\s*k\b", limpo, flags=re.IGNORECASE)
    if m:
        num_part     = m.group(1)
        multiplicador = 1_000.0
    else:
        m = re.search(r"[0-9.,]+", limpo)
        if not m:
            return INVALIDO
        num_part = m.group(0)

    n = _parse_preco_numerico(num_part)
    if n is None:
        return INVALIDO
    return f"{n * multiplicador:.2f}"


# ---------------------------------------------------------------------------
# Higienização da quilometragem
# ---------------------------------------------------------------------------
_ZERO_ALIASES = {
    "new", "novo", "new car", "carro novo",
    "zero miles", "zero milhas", "zero",
    "0 mi", "0 miles", "0mi", "0 km",
}


def sanitizar_quilometragem(valor: Any) -> str:
    if valor is None:
        return INVALIDO

    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        return str(int(round(float(valor))))

    s = str(valor).strip()
    if not s:
        return INVALIDO

    baixo = s.lower()
    if baixo in _ZERO_ALIASES:
        return "0"

    eh_km = bool(re.search(r"\bkm\b|\bkilom", baixo))

    # Texto puro sem dígitos
    if not re.search(r"\d", s):
        textual = re.sub(
            r"\b(mi|miles?|km|kilometers?|kilometres?|milhas?)\b", " ", baixo
        )
        n = _parse_preco_textual(textual.strip())
        if n is None:
            return INVALIDO
        milhas = n * (0.621371 if eh_km else 1.0)
        return str(int(round(milhas)))

    m = re.search(r"[0-9][0-9.,]*", s)
    if not m:
        return INVALIDO
    n = _parse_preco_numerico(m.group(0))
    if n is None:
        return INVALIDO
    if eh_km:
        n = n * 0.621371
    return str(int(round(n)))


# ---------------------------------------------------------------------------
# Higienização do método de pagamento
# ---------------------------------------------------------------------------
def _normalizar_chave(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[\W_]+", " ", s.lower()).strip()
    return re.sub(r"\s+", " ", s)


PAGAMENTO_LOOKUP: dict[str, str] = {
    # Cartão de Crédito
    "cartao de credito":        "Cartão de Crédito",
    "credito":                  "Cartão de Crédito",
    "credit card":              "Cartão de Crédito",
    "credit":                   "Cartão de Crédito",
    "creditcard":               "Cartão de Crédito",
    "credit card payment":      "Cartão de Crédito",
    # Cartão de Débito
    "cartao de debito":         "Cartão de Débito",
    "debit card":               "Cartão de Débito",
    "debitcard":                "Cartão de Débito",
    # Transferência Bancária
    "transferencia bancaria":   "Transferência Bancária",
    "bank transfer":            "Transferência Bancária",
    "bank":                     "Transferência Bancária",
    "bank transfer":            "Transferência Bancária",
    "bank transfer":            "Transferência Bancária",
    "bank transfer":            "Transferência Bancária",
    # Wire / Transferência Eletrônica
    "wire transfer":            "Transferência Eletrônica",
    "wire":                     "Transferência Eletrônica",
    "wiretransfer":             "Transferência Eletrônica",
    "wire transfer":            "Transferência Eletrônica",
    "bank wire":                "Transferência Eletrônica",
    # Financiamento
    "financiamento":            "Financiamento",
    "financing":                "Financiamento",
    "finance":                  "Financiamento",
    "financing plan":           "Financiamento",
    # Leasing / Arrendamento
    "leasing":                  "Leasing",
    "lease":                    "Leasing",
    "lease plan":               "Leasing",
    # Dinheiro
    "dinheiro":                 "Dinheiro",
    "cash":                     "Dinheiro",
    "cash payment":             "Dinheiro",
    # ACH
    "ach":                      "Pagamento ACH",
    "ach payment":              "Pagamento ACH",
    # Criptomoedas
    "cripto":                   "Pagamento em Criptomoedas",
    "crypto":                   "Pagamento em Criptomoedas",
    "crypto payment":           "Pagamento em Criptomoedas",
    "criptomoeda":              "Pagamento em Criptomoedas",
    "criptomoedas":             "Pagamento em Criptomoedas",
}


def sanitizar_pagamento(valor: Any) -> str:
    if valor is None:
        return INVALIDO
    s = str(valor).strip()
    if not s:
        return INVALIDO
    chave = _normalizar_chave(s)
    if chave in PAGAMENTO_LOOKUP:
        return PAGAMENTO_LOOKUP[chave]
    return " ".join(w.capitalize() for w in chave.split())


# ---------------------------------------------------------------------------
# Higienização da cidade
# ---------------------------------------------------------------------------
def sanitizar_cidade(valor: Any) -> str:
    if valor is None:
        return INVALIDO
    s = " ".join(str(valor).split()).strip()
    if not s:
        return INVALIDO

    saida: list[str] = []
    for p in s.split():
        low = p.lower().rstrip(".")
        if low == "st":
            saida.append("St.")
        elif low == "mt":
            saida.append("Mt.")
        elif "-" in p:
            saida.append("-".join(seg.capitalize() for seg in p.split("-")))
        else:
            saida.append(p.capitalize())
    return " ".join(saida)


# ---------------------------------------------------------------------------
# Higienização do estado (EUA)
# ---------------------------------------------------------------------------
ESTADOS_EUA: dict[str, str] = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR",
    "california": "CA", "colorado": "CO", "connecticut": "CT", "delaware": "DE",
    "florida": "FL", "georgia": "GA", "hawaii": "HI", "idaho": "ID",
    "illinois": "IL", "indiana": "IN", "iowa": "IA", "kansas": "KS",
    "kentucky": "KY", "louisiana": "LA", "maine": "ME", "maryland": "MD",
    "massachusetts": "MA", "michigan": "MI", "minnesota": "MN",
    "mississippi": "MS", "missouri": "MO", "montana": "MT", "nebraska": "NE",
    "nevada": "NV", "new hampshire": "NH", "new jersey": "NJ",
    "new mexico": "NM", "new york": "NY", "north carolina": "NC",
    "north dakota": "ND", "ohio": "OH", "oklahoma": "OK", "oregon": "OR",
    "pennsylvania": "PA", "rhode island": "RI", "south carolina": "SC",
    "south dakota": "SD", "tennessee": "TN", "texas": "TX", "utah": "UT",
    "vermont": "VT", "virginia": "VA", "washington": "WA",
    "west virginia": "WV", "wisconsin": "WI", "wyoming": "WY",
    "district of columbia": "DC",
}
CODIGOS_ESTADOS = set(ESTADOS_EUA.values())


def sanitizar_estado(valor: Any) -> str:
    if valor is None:
        return INVALIDO
    s = str(valor).strip()
    if not s:
        return INVALIDO
    maiusculo = s.upper().strip()
    if maiusculo in CODIGOS_ESTADOS:
        return maiusculo
    chave = " ".join(s.lower().split())
    return ESTADOS_EUA.get(chave, INVALIDO)


# ---------------------------------------------------------------------------
# Higienização do status de entrega
# ---------------------------------------------------------------------------
ENTREGA_LOOKUP: dict[str, str] = {
    # Entregue
    "delivered":          "Entregue",
    "deliverd":           "Entregue",   # erro ortográfico comum
    "entregue":           "Entregue",
    # Pendente
    "pending":            "Pendente",
    "pendente":           "Pendente",
    # Em Trânsito
    "in transit":         "Em Trânsito",
    "in-transit":         "Em Trânsito",
    "em transito":        "Em Trânsito",
    "em transito":        "Em Trânsito",
    # Cancelado
    "cancelled":          "Cancelado",
    "canceled":           "Cancelado",
    "cancelado":          "Cancelado",
    # Aguardando Entrega
    "awaiting delivery":  "Aguardando Entrega",
    "aguardando entrega": "Aguardando Entrega",
    # Aguardando Recolhimento
    "awaiting pickup":    "Aguardando Recolhimento",
    # Aprovação Pendente
    "pending approval":   "Aprovação Pendente",
    # Enviado
    "shipped":            "Enviado",
    # Revisão Pendente
    "pending review":     "Revisão Pendente",
    "awaiting review":    "Aguardando Revisão",
}


def sanitizar_entrega(valor: Any) -> str:
    if valor is None:
        return INVALIDO
    s = str(valor).strip()
    if not s:
        return INVALIDO
    chave = _normalizar_chave(s)
    if chave in ENTREGA_LOOKUP:
        return ENTREGA_LOOKUP[chave]
    resultado = " ".join(w.capitalize() for w in chave.split())
    return resultado if resultado else INVALIDO


# ---------------------------------------------------------------------------
# Pipeline de colunas
# (coluna_origem, nome_coluna_sanitizada | None, função_sanitizadora | None)
# ---------------------------------------------------------------------------
PIPELINE_COLUNAS = [
    ("sale_id",          None,                      None),
    ("sale_date",        "SaleDateSanitized",       sanitizar_data),
    ("customer_name",    None,                      None),
    ("porsche_model",    "PorscheModelSanitized",   sanitizar_modelo),
    ("model_year",       "ModelYearSanitized",      sanitizar_ano),
    ("sale_price",       "SalesPriceSanitized",     sanitizar_preco),
    ("vehicle_mileage",  "VehicleMileageSanitized", sanitizar_quilometragem),
    ("payment_method",   "PayMethodSanitized",      sanitizar_pagamento),
    ("city",             "CitySanitized",           sanitizar_cidade),
    ("state",            "StateSanitized",          sanitizar_estado),
    ("salesperson",      None,                      None),
    ("delivery_status",  "DeliveryStatusSanitized", sanitizar_entrega),
]


# ---------------------------------------------------------------------------
# Processamento da pasta de trabalho
# ---------------------------------------------------------------------------
def processar_pasta_de_trabalho(
    caminho_entrada: Path, caminho_saida: Path
) -> dict[str, int]:
    wb_entrada = load_workbook(caminho_entrada, data_only=True)
    ws_entrada = wb_entrada.active

    linhas = list(ws_entrada.iter_rows(values_only=True))
    if not linhas:
        raise ValueError("Pasta de trabalho vazia")

    cabecalho    = [str(h).strip() if h is not None else "" for h in linhas[0]]
    nome_para_idx = {h: i for i, h in enumerate(cabecalho)}

    ausentes = [c for c, _, _ in PIPELINE_COLUNAS if c not in nome_para_idx]
    if ausentes:
        raise ValueError(f"Colunas obrigatórias ausentes: {ausentes}")

    wb_saida = Workbook()
    ws_saida = wb_saida.active
    ws_saida.title = "Sanitizado"

    # Cabeçalho de saída
    cabecalho_saida: list[str] = []
    for src, san_name, _ in PIPELINE_COLUNAS:
        cabecalho_saida.append(src)
        if san_name:
            cabecalho_saida.append(san_name)
    ws_saida.append(cabecalho_saida)

    # Estilo do cabeçalho
    fonte_negrito    = Font(bold=True, color="FFFFFF")
    preench_orig     = PatternFill("solid", fgColor="1F2937")  # cinza-escuro
    preench_sanitiz  = PatternFill("solid", fgColor="065F46")  # verde-escuro
    for col_idx, nome in enumerate(cabecalho_saida, start=1):
        cel = ws_saida.cell(row=1, column=col_idx)
        cel.font  = fonte_negrito
        cel.fill  = preench_sanitiz if nome.endswith("Sanitized") else preench_orig

    # Dados
    contagem_invalidos: dict[str, int] = {}
    for r in linhas[1:]:
        linha_saida: list[Any] = []
        for src, san_name, fn in PIPELINE_COLUNAS:
            bruto = r[nome_para_idx[src]]
            linha_saida.append(bruto)
            if san_name and fn is not None:
                limpo = fn(bruto)
                if limpo == INVALIDO:
                    contagem_invalidos[san_name] = contagem_invalidos.get(san_name, 0) + 1
                linha_saida.append(limpo)
        ws_saida.append(linha_saida)

    # Largura das colunas
    for col_idx in range(1, len(cabecalho_saida) + 1):
        letra = get_column_letter(col_idx)
        ws_saida.column_dimensions[letra].width = max(14, len(cabecalho_saida[col_idx - 1]) + 2)
    ws_saida.freeze_panes = "A2"

    wb_saida.save(caminho_saida)
    return contagem_invalidos


# ---------------------------------------------------------------------------
# Entrada principal
# ---------------------------------------------------------------------------
def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("entrada", nargs="?", default=str(INPUT_PATH))
    parser.add_argument("saida",   nargs="?", default=str(OUTPUT_PATH))
    args = parser.parse_args(argv)

    caminho_entrada = Path(args.entrada)
    caminho_saida   = Path(args.saida)

    if not caminho_entrada.exists():
        print(f"ERRO: arquivo não encontrado: {caminho_entrada}", file=sys.stderr)
        return 1

    stats = processar_pasta_de_trabalho(caminho_entrada, caminho_saida)
    print(f"OK -> {caminho_saida}")
    if stats:
        print("Contagem de INVALIDOS por coluna sanitizada:")
        for k, v in sorted(stats.items()):
            print(f"  {k}: {v}")
    else:
        print("Nenhum valor inválido encontrado.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
